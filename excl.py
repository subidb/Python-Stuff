import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Color, PatternFill, Font, Border

wb = xl.load_workbook("goods.xlsx")
sheet = wb["Sheet1"]
ws = wb.active

redFill = PatternFill(start_color='FFFF0000',
                      end_color='FFFF0000',
                      fill_type='solid')


print(help(PatternFill))
def getcellno(cel):
    cell_str = str(cell)
    cell_n = cell_str[15:17]
    return cell_n


for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row,3)
    cll = getcellno(cell)
    if cell.value > 200:
        # cll = getcellno(cell)
        # ws[cll].fill = redFill
        for c in range(1, sheet.max_column + 1):
            c = sheet.cell(row, c)
            c.fill = PatternFill(fgColor="FFFF0000", fill_type="solid")



values = Reference(sheet, min_row = 2, max_row = sheet.max_row, min_col = 4, max_col = 4)

chart = BarChart()
chart.add_data(values)
# sheet.add_chart(chart, "c2")
wb.save("goods3.xlsx")
