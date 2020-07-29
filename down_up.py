import os
from imbox import Imbox 
import traceback
from datetime import datetime
import smtplib, ssl
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders
from email.message import EmailMessage
import requests
import openpyxl as xl
from openpyxl.styles import Color, PatternFill, Font, Border



host = "imap.gmail.com"
sender_gmail = "subid@isarlend.com"
password = input("Enter your password: ")



def download_coba():
    download_folder = "."
    if not os.path.isdir(download_folder):
        os.makedirs(download_folder, exist_ok=True)
        
    mail = Imbox(host, username=sender_gmail, password=password, ssl=True, ssl_context=None, starttls=False)
    messages = mail.messages(sent_from='alexander@isarlend.com', date__on = datetime.now()) # filters below

    for (uid, message) in messages:
        mail.mark_seen(uid) # optional, mark message as read

        for idx, attachment in enumerate(message.attachments):
            try:
                att_fn = attachment.get('filename')
                download_path = f"{download_folder}/{att_fn}"
                print(download_path)
                with open(download_path, "wb") as fp:
                    fp.write(attachment.get('content').read())
                    print("downloaded coba file!\n")
            except:
                #pass
                # print(traceback.print_exc())
                print("error downloading")

    mail.logout()


def edit_xl():
    path = '.'
    files = os.listdir(path)
    for index, file in enumerate(files):
        if os.path.splitext(file)[1] == '.xlsx':
            os.rename(os.path.join(path, file), os.path.join(path, ''.join(["coba", '.xlsx'])))
    wb = xl.load_workbook("coba.xlsx")
    sheet = wb["Umsätze"]
    # ws = wb.active
    for row in range(3, sheet.max_row + 1):
        cell = sheet.cell(row,8)
        if cell.value[0] == '-':
            newval = '-' + "€" + cell.value[1:]
        else:
            newval = "€" + cell.value[0:]
        newval = newval.replace(".", '^').replace(",", '.')
        cell.value = newval.replace("^", ',')
    wb.save("coba_update.xlsx")
    
def send_email(recipient_email, subject, content, attachment_file):
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = sender_gmail
    msg['To'] = recipient_email
    msg.set_content(content)

    with open(attachment_file, 'rb') as f:
        file_data = f.read()
    msg.add_attachment(file_data, maintype="application", subtype="csv", filename=attachment_file)

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(sender_gmail, password)
        smtp.send_message(msg)

if __name__ == "__main__":

    download_coba()
    edit_xl()
    receiver_email = "mehmet@fulfin.io"
    subject = "Coba update " + str(datetime.today().strftime('%Y-%m-%d'))
    attachment_file = 'coba_update.xlsx'
    send_email(receiver_email, subject, "", attachment_file)
    print("\nEmail Sent!")
